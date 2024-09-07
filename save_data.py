import openpyxl
from openpyxl.styles import Font

# Función para crear o actualizar el archivo Excel con los datos del ticket
def save(ticket_data:list, fecha:str, path:str="tickets.xlsx"):
    try:
        # Intentamos abrir el archivo existente
        workbook = openpyxl.load_workbook(path)
    except FileNotFoundError:
        # Si no existe el archivo, lo creamos
        workbook = openpyxl.Workbook()

    # Seleccionamos la hoja activa (o creamos una nueva)
    sheet = workbook.active
    sheet.title = "Tickets"

    # Encontramos la última fila ocupada
    ultima_fila = sheet.max_row

    # Si es la primera vez, escribimos los encabezados
    if ultima_fila == 1:
        encabezados = ['fecha', 'producto', 'cantidad', 'precio unitario', 'precio total', 'precio total del ticket']
        for col_num, encabezado in enumerate(encabezados, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = encabezado
            cell.font = Font(bold=True)

    # Escribimos la fecha en la primera celda de la fila
    sheet.cell(row=ultima_fila + 1, column=1).value = fecha

    # Agregamos los datos del ticket
    total_precio_ticket = 0
    for i, producto in enumerate(ticket_data, start=0):
        # Colocamos el producto en la fila actual
        sheet.cell(row=ultima_fila + 1 + i, column=2).value = producto[0]  # Nombre del producto
        sheet.cell(row=ultima_fila + 1 + i, column=3).value = producto[1]  # Cantidad
        sheet.cell(row=ultima_fila + 1 + i, column=4).value = producto[2]  # Precio unitario
        sheet.cell(row=ultima_fila + 1 + i, column=5).value = producto[3]  # Precio total

        # Eliminamos el símbolo de moneda y convertimos el total a float
        precio_total_limpio = float(producto[3].replace("€", "").replace(",", ".").strip())
        
        # Acumulamos el total del ticket
        total_precio_ticket += precio_total_limpio

    # Escribimos el total del ticket en la última celda de la primera fila del ticket
    sheet.cell(row=ultima_fila + 1, column=6).value = total_precio_ticket

    # Guardamos el archivo
    workbook.save(path)
    print(f"Los datos del ticket se han guardado en {path}.")

